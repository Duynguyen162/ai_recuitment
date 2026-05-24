import asyncio
import os
import sys
from fastapi import APIRouter, File, HTTPException
from sqlalchemy.orm import Session
from app.core.enum import RoleEnum
from app.crud.crud_candidate_detail import create_candidate_certification, create_candidate_cv, create_candidate_education, create_candidate_experience, delete_candidate_certification, delete_candidate_cv, delete_candidate_education, delete_candidate_experience, get_candidate_certification, get_candidate_cv, get_candidate_cv_by_id, get_candidate_education, get_candidate_experience, get_full_candidate_cv, update_candidate_certification, update_candidate_education, update_candidate_experience
from app.db.database import get_db
from app.models.user import User
from app.models.candidate_profiles import CandidateProfile
from app.models.candidate_details import CandidateExperience , CandidateEducation, CandidateCertification, CVUpload
from app.api.deps import get_current_user
from app.schemas.candidate_details_schema import CandidateExperienceCreate ,CandidateExperienceResponse
from app.schemas.candidate_details_schema import CandidateEducationCreate, CandidateEducationResponse
from app.schemas.candidate_profiles_schema import CandidateProfileResponse , CandidateProfileUpdate
from app.schemas.candidate_details_schema import CandidateCertificationCreate, CandidateCertificationResponse
from app.schemas.candidate_details_schema import CVUploadCreate, CVUploadResponse
from app.crud.crud_candidate_profile import get_candidate_profile, upsert_profile
from fastapi.responses import FileResponse
from app.schemas.base_schema import ResponseSchema
from app.api.deps import get_current_candidate_profile
from fastapi import UploadFile, File   
from sqlalchemy.orm import Session
from fastapi import APIRouter, Depends, Response
from fastapi.responses import Response
from jinja2 import Environment, FileSystemLoader
from playwright.async_api import async_playwright
from urllib.parse import quote

router = APIRouter(tags=["Candidate Profile"])

@router.get("/profileCandidate", response_model = ResponseSchema[CandidateProfileResponse])
def get_profile(db: Session = Depends(get_db), 
                current_user: User = Depends(get_current_user),
                profile: CandidateProfile = Depends(get_current_candidate_profile)
               ): 
    # Chuyển đổi từ SQLAlchemy model ( 1 đối tượng) sang Pydantic model(kiểu json)
    profile_response = CandidateProfileResponse.model_validate(profile)

    return ResponseSchema(
        success=True,
        data=profile_response,
        error=None,
        meta=None
    )

@router.post("/profileCandidate", response_model = ResponseSchema[CandidateProfileResponse])
def update_profile(profile_in: CandidateProfileUpdate,
                db: Session = Depends(get_db), 
                current_user: User = Depends(get_current_user),
                profile: CandidateProfile = Depends(get_current_candidate_profile)
                ):   
    upsert_profile(db, current_user.id, profile_in)
    profile_response = CandidateProfileResponse.model_validate(profile)
    return ResponseSchema(
        success=True,
        data=profile_response,
        error=None,
        meta=None
    )
#=============================================================#
#====================== KINH NGHIỆM ==========================#
#=============================================================#
@router.get("/experiences",response_model = ResponseSchema[list[CandidateExperienceResponse]])
def get_experience(db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user),
                       profile: CandidateProfile = Depends(get_current_candidate_profile)):
    exp_detail = get_candidate_experience(db,profile)
    return ResponseSchema(
        success=True,
        data=exp_detail,
        error=None,
        meta=None
    )

@router.post("/experiences", response_model = ResponseSchema[CandidateExperienceResponse])
def create_experience(experience_in: CandidateExperienceCreate,
                       db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user),
                       profile: CandidateProfile = Depends(get_current_candidate_profile)
                       ):
    candidate_exp = create_candidate_experience(db, experience_in.model_dump(), profile.id) # truyền candidate_id vào để liên kết
    return ResponseSchema(
        success=True,
        data=CandidateExperienceResponse.model_validate(candidate_exp),
        error=None,
        meta=None
    )

@router.put("/experiences/{experience_id}", 
            response_model = ResponseSchema[CandidateExperienceResponse])
def update_experience(experience_id: int,
                    experience_in: CandidateExperienceCreate,
                    db: Session = Depends(get_db),
                    current_user: User = Depends(get_current_user),
                    profile: CandidateProfile = Depends(get_current_candidate_profile)
                       ):
    exp = update_candidate_experience(db, experience_id, experience_in.model_dump(), current_user) # truyền dữ liệu đã chuyển đổi thành dict
    return ResponseSchema(
        success=True,
        data=CandidateExperienceResponse.model_validate(exp),
        error=None,
        meta=None
    )

@router.delete("/experiences/{experience_id}")
def delete_experience(experience_id: int,
                       db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user),
                       profile: CandidateProfile = Depends(get_current_candidate_profile)
                       ):
    delete_candidate_experience(db, experience_id, current_user )
    return ResponseSchema(
        success=True,
        data=None,
        error=None,
        meta=None
    )
#=============================================================#
#====================== HỌC VẤN ==============================#
#=============================================================#
@router.get("/educations",response_model = ResponseSchema[list[CandidateEducationResponse]])
def get_education(db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user),
                       profile: CandidateProfile = Depends(get_current_candidate_profile)):
    edu_detail = get_candidate_education(db,profile)
    return ResponseSchema(
        success=True,
        data=edu_detail,
        error=None,
        meta=None
    )

@router.post("/educations", response_model = ResponseSchema[CandidateEducationResponse])
def create_education(education_in: CandidateEducationCreate,
                     db: Session = Depends(get_db),
                     current_user: User = Depends(get_current_user),
                     profile: CandidateProfile = Depends(get_current_candidate_profile)
                     ): 
    education = create_candidate_education(db, education_in.model_dump(), profile.id) # truyền candidate_id vào để liên kết
    return ResponseSchema(
        success=True,
        data=CandidateEducationResponse.model_validate(education),
        error=None,
        meta=None
    )
@router.put("/educations/{education_id}", response_model = ResponseSchema[CandidateEducationResponse])
def update_education(education_id: int,
                    education_in: CandidateEducationCreate,
                    db: Session = Depends(get_db),
                    current_user: User = Depends(get_current_user),
                    profile: CandidateProfile = Depends(get_current_candidate_profile)
                       ):
    education = update_candidate_education(db, education_id, education_in.model_dump(), current_user)
    return ResponseSchema(
        success=True,
        data=CandidateEducationResponse.model_validate(education),
        error=None,
        meta=None
    )
@router.delete("/educations/{education_id}")
def delete_education(education_id: int,
                     db: Session = Depends(get_db),
                     current_user: User = Depends(get_current_user),
                     profile: CandidateProfile = Depends(get_current_candidate_profile)
                     ):
    delete_candidate_education(db, education_id, current_user)
    return ResponseSchema(
        success=True,
        data=None,
        error=None,
        meta=None
    )

#=============================================================#
#====================== BẰNG CẤP ==============================#
#=============================================================#
@router.get("/certifications",response_model = ResponseSchema[list[CandidateCertificationResponse]])
def get_certification(db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user),
                       profile: CandidateProfile = Depends(get_current_candidate_profile)):
    exp_detail = get_candidate_certification(db,profile)
    return ResponseSchema(
        success=True,
        data=exp_detail,
        error=None,
        meta=None
    )

@router.post("/certifications", response_model = ResponseSchema[CandidateCertificationResponse])
def create_certification(certification_in: CandidateCertificationCreate,
                         db: Session = Depends(get_db),
                         current_user: User = Depends(get_current_user),
                         profile: CandidateProfile = Depends(get_current_candidate_profile)
                         ):  
    certification = create_candidate_certification(db, certification_in.model_dump(), profile.id) # truyền candidate_id vào để liên kết
    return ResponseSchema(
        success=True,
        data=CandidateCertificationResponse.model_validate(certification),
        error=None,
        meta=None
    )
@router.put("/certifications/{certification_id}", response_model = ResponseSchema[CandidateCertificationResponse])
def update_certification(certification_id: int,
                         certification_in: CandidateCertificationCreate,
                         db: Session = Depends(get_db),
                         current_user: User = Depends(get_current_user),
                         profile: CandidateProfile = Depends(get_current_candidate_profile)
                       ):
    
    certification = update_candidate_certification(db, certification_id, certification_in.model_dump(), current_user)
    return ResponseSchema(
        success=True,
        data=CandidateCertificationResponse.model_validate(certification),
        error=None,
        meta=None
    )
@router.delete("/certifications/{certification_id}")
def delete_certification(certification_id: int,
                         db: Session = Depends(get_db),
                         current_user: User = Depends(get_current_user),
                         profile: CandidateProfile = Depends(get_current_candidate_profile)
                     ):
    delete_candidate_certification(db, certification_id, current_user)
    return ResponseSchema(
        success=True,
        data=None,
        error=None,
        meta=None
    )
#=============================================================#
#=========================== CV ==============================#
#=============================================================#


from app.core.config import settings

@router.get("/cv_upload",response_model = ResponseSchema[list[CVUploadResponse]])
def get_list_cv(db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user),
                       profile: CandidateProfile = Depends(get_current_candidate_profile)):
    """lấy danh sách các cv """
    cvs = get_candidate_cv(db,profile)
    
    data = [
        CVUploadResponse(
            id=cv.id,
            candidate_id=cv.candidate_id,
            file_name=cv.file_name,
            file_url=f"{settings.BASE_URL}/{cv.file_url}" if cv.file_url else None
        )
        for cv in cvs
    ]

    return ResponseSchema(
        success=True,
        data=data,
        error=None,
        meta=None
    )

@router.get("/cv_upload/{cv_id}/view")
def view_cv_file(
    cv_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    profile: CandidateProfile = Depends(get_current_candidate_profile)
):
    cv = get_candidate_cv_by_id(db, profile, cv_id)
    file_path = cv.file_url 
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File không tồn tại trên server")

    return FileResponse(
        path=file_path, 
        filename=cv.file_name,
        # Bật Expose Headers để Frontend đọc được định dạng file
        headers={"Access-Control-Expose-Headers": "Content-Disposition"} 
    )

async def generate_pdf_from_html(html: str):
    """Convert HTML to PDF using Playwright"""
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    
    browser = None
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            
            # Set viewport để layout đúng
            await page.set_viewport_size({"width": 1200, "height": 1600})
            
            await page.set_content(html, wait_until="networkidle")
            
            pdf = await page.pdf(
                format="A4",
                print_background=True,  
                margin={"top": "10mm", "right": "10mm", "bottom": "10mm", "left": "10mm"},
                display_header_footer=False
            )
            await browser.close()
            return pdf
    except Exception as e:
        if browser:
            await browser.close()
        raise HTTPException(status_code=500, detail=f"Lỗi tạo PDF: {str(e)}")


@router.get("/export_cv")
async def export_cv_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    profile: CandidateProfile = Depends(get_current_candidate_profile)
):
    """ api export ra cv theo fomat của hệ thống"""
    full_profile = get_full_candidate_cv(db, profile.id)
    
    if not full_profile:
        raise HTTPException(status_code=404, detail="Không tìm thấy CV")

    context = {
        "full_name": full_profile.full_name or "Unknown",
        "phone": full_profile.phone or "",
        "email": current_user.email,
        "bio": full_profile.bio or "",
        "avatar_url": full_profile.avatar_url,

        "links": {
            "github": full_profile.github_url or "",
            "linkedin": full_profile.linkedin_url or "",
            "portfolio": full_profile.portfolio_url or "",
        },

        "skill_tags": full_profile.skill_tags or [],
        "years_of_experience": full_profile.years_of_experience or 0,

        "experiences": [
            {
                "company_name": exp.company_name,
                "job_title": exp.job_title, 
                "description": exp.description,
            }
            for exp in (full_profile.experiences or [])
        ],

        "educations": [
            {
                "institution_name": edu.institution_name,
                "major": edu.major,
                "degree": edu.degree,
            }
            for edu in (full_profile.educations or [])
        ],

        "certifications": [
            {
                "name": cert.name,
                "issuer": cert.issuer,
            }
            for cert in (full_profile.certifications or [])
        ],
    }

    try:
        # Render HTML
        env = Environment(loader=FileSystemLoader("app/templates"))
        template = env.get_template("cv_template.html")
        html_out = template.render(context)

        # Convert to PDF
        pdf_bytes = await generate_pdf_from_html(html_out)
        
        filename = f"CV_{full_profile.full_name}.pdf"
        safe_filename = quote(filename, safe='')
        
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename*=UTF-8\'\'{safe_filename}'
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi xuất CV: {str(e)}")

@router.post("/cv_upload", response_model = ResponseSchema[CVUploadResponse])
def create_cv_upload(file: UploadFile = File(...),
                     db: Session = Depends(get_db),
                     current_user: User = Depends(get_current_user),
                     profile: CandidateProfile = Depends(get_current_candidate_profile)
                     ):
    cv_record = create_candidate_cv(db, file, profile.id, profile.user_id)
    return ResponseSchema(
        success=True,
        data=CVUploadResponse.model_validate(cv_record),
        error=None,
        meta=None
    )

@router.delete("/cv_upload/{cv_upload_id}")
def delete_cv_upload(cv_upload_id: int,
                         db: Session = Depends(get_db),
                         current_user: User = Depends(get_current_user),
                         profile: CandidateProfile = Depends(get_current_candidate_profile)
                     ):
    delete_candidate_cv(db,cv_upload_id,current_user)
    return ResponseSchema(
        success=True,
        data=None,
        error=None,
        meta=None
    )


def generate_default_excel_template(file_path: str):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    
    # Định dạng chung cho Excel
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Màu xanh dương đậm
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 1. Sheet Thông tin cá nhân
    ws_info = wb.active
    ws_info.title = "Thông tin cá nhân"
    headers_info = [
        "Họ và tên (*)",
        "Số điện thoại (*)",
        "Giới thiệu ngắn",
        "Portfolio URL",
        "LinkedIn URL",
        "GitHub URL",
        "Kỹ năng (Ngăn cách bởi dấu phẩy)",
        "Số năm kinh nghiệm"
    ]
    ws_info.append(headers_info)
    ws_info.append([
        "Nguyễn Văn A",
        "0987654321",
        "Tôi là lập trình viên backend có kinh nghiệm làm việc với FastAPI, Python và SQL.",
        "https://myportfolio.com",
        "https://linkedin.com/in/nguyenvana",
        "https://github.com/nguyenvana",
        "Python, FastAPI, SQL, Docker, ReactJS",
        3
    ])
    
    # 2. Sheet Kinh nghiệm làm việc
    ws_exp = wb.create_sheet(title="Kinh nghiệm làm việc")
    headers_exp = ["Tên công ty (*)", "Chức danh / Vị trí (*)", "Mô tả công việc"]
    ws_exp.append(headers_exp)
    ws_exp.append([
        "Công ty Công nghệ ABC",
        "Backend Developer",
        "Thiết kế cơ sở dữ liệu, phát triển API và tối ưu hiệu năng hệ thống bằng FastAPI."
    ])
    ws_exp.append([
        "Tập đoàn XYZ",
        "Python Developer Intern",
        "Viết các script tự động hóa, kiểm thử API và hỗ trợ bảo trì hệ thống backend."
    ])
    
    # 3. Sheet Học vấn
    ws_edu = wb.create_sheet(title="Học vấn")
    headers_edu = ["Tên trường / Trung tâm (*)", "Chuyên ngành (*)", "Bằng cấp (*) (Ví dụ: Cử nhân, Kỹ sư...)"]
    ws_edu.append(headers_edu)
    ws_edu.append([
        "Đại học Bách Khoa Hà Nội",
        "Khoa học Máy tính",
        "Kỹ sư"
    ])
    
    # 4. Sheet Chứng chỉ
    ws_cert = wb.create_sheet(title="Chứng chỉ")
    headers_cert = ["Tên chứng chỉ (*)", "Tổ chức cấp (*)"]
    ws_cert.append(headers_cert)
    ws_cert.append([
        "AWS Certified Solutions Architect",
        "Amazon Web Services"
    ])
    
    # Định dạng giao diện cho tất cả các sheets
    for ws in [ws_info, ws_exp, ws_edu, ws_cert]:
        ws.row_dimensions[1].height = 28
        # Định dạng Header
        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
        # Định dạng hàng dữ liệu mẫu
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 22
            for cell in ws[row]:
                cell.alignment = align_left
        # Tự động căn chỉnh độ rộng cột
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                val_len = len(val.encode('utf-8'))
                if val_len > max_len:
                    max_len = val_len
            col_letter = get_column_letter(col[0].column)
            # Giới hạn độ rộng cột trong khoảng [15, 50]
            ws.column_dimensions[col_letter].width = max(min(max_len // 2 + 5, 50), 15)
            
    # Lưu file
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)


@router.get("/profileCandidate/template")
def download_profile_template():
    """Tải file Excel mẫu để điền thông tin hồ sơ"""
    template_path = os.path.join("app", "templates", "candidate_profile_template.xlsx")
    
    # Nếu chưa có file template tĩnh, tự động tạo ra một file chất lượng cao để tải về
    if not os.path.exists(template_path):
        try:
            generate_default_excel_template(template_path)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Không thể tạo file template: {str(e)}")
            
    return FileResponse(
        path=template_path,
        filename="candidate_profile_template.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Access-Control-Expose-Headers": "Content-Disposition"}
    )

def is_instruction_row(row_values):
    text = " ".join([str(v or "").lower() for v in row_values])
    
    keywords = [
        "hàng màu vàng",
        "dữ liệu mẫu",
        "hãy xoá",
        "bắt buộc"
    ]
    
    return any(k in text for k in keywords)
    
@router.post("/profileCandidate/import", response_model=ResponseSchema[CandidateProfileResponse])
def import_profile_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    profile: CandidateProfile = Depends(get_current_candidate_profile)
):
    """Nhập dữ liệu hồ sơ từ file Excel (Thông tin cá nhân, Kinh nghiệm, Học vấn, Chứng chỉ)"""
    import openpyxl
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Định dạng file không hỗ trợ, vui lòng sử dụng file .xlsx")
        
    try:
        # Load workbook từ file tải lên
        wb = openpyxl.load_workbook(file.file)
        
        # 1. Đọc sheet "Thông tin cá nhân"
        if "Thông tin cá nhân" in wb.sheetnames:
            ws_info = wb["Thông tin cá nhân"]
            
            # Tìm dòng chứa dữ liệu đầu tiên (bỏ qua các dòng hướng dẫn)
            target_row = None
            for r in range(2, ws_info.max_row + 1):
                row_vals = [ws_info.cell(row=r, column=col).value for col in range(1, 9)]
                if any(v is not None for v in row_vals) and not is_instruction_row(row_vals):
                    target_row = r
                    break
            
            if target_row:
                full_name = ws_info.cell(row=target_row, column=1).value
                phone = ws_info.cell(row=target_row, column=2).value
                bio = ws_info.cell(row=target_row, column=3).value
                portfolio_url = ws_info.cell(row=target_row, column=4).value
                linkedin_url = ws_info.cell(row=target_row, column=5).value
                github_url = ws_info.cell(row=target_row, column=6).value
                skills_str = ws_info.cell(row=target_row, column=7).value
                years_exp = ws_info.cell(row=target_row, column=8).value
                
                # Cập nhật thông tin cơ bản
                if full_name is not None:
                    profile.full_name = str(full_name).strip()
                if phone is not None:
                    profile.phone = str(phone).strip()
                if bio is not None:
                    profile.bio = str(bio).strip()
                if portfolio_url is not None:
                    profile.portfolio_url = str(portfolio_url).strip()
                if linkedin_url is not None:
                    profile.linkedin_url = str(linkedin_url).strip()
                if github_url is not None:
                    profile.github_url = str(github_url).strip()
                    
                if years_exp is not None:
                    try:
                        profile.years_of_experience = int(years_exp)
                    except ValueError:
                        pass
                        
                if skills_str is not None:
                    profile.skill_tags = [s.strip() for s in str(skills_str).split(",") if s.strip()]
                    
                db.add(profile)

        # 2. Đọc sheet "Kinh nghiệm làm việc"
        if "Kinh nghiệm làm việc" in wb.sheetnames:
            ws_exp = wb["Kinh nghiệm làm việc"]
            # Xóa các kinh nghiệm cũ
            db.query(CandidateExperience).filter(CandidateExperience.candidate_id == profile.id).delete()
            # Thêm các kinh nghiệm mới từ file Excel
            for r in range(2, ws_exp.max_row + 1):
                row_vals = [ws_exp.cell(row=r, column=col).value for col in range(1, 4)]
                if is_instruction_row(row_vals):
                    continue
                    
                comp_name = ws_exp.cell(row=r, column=1).value
                job_title = ws_exp.cell(row=r, column=2).value
                desc = ws_exp.cell(row=r, column=3).value
                
                # Chỉ lưu dòng có nhập thông tin công ty hoặc vị trí
                if comp_name or job_title:
                    db_exp = CandidateExperience(
                        candidate_id=profile.id,
                        company_name=str(comp_name or "").strip(),
                        job_title=str(job_title or "").strip(),
                        description=str(desc or "").strip()
                    )
                    db.add(db_exp)

        # 3. Đọc sheet "Học vấn"
        if "Học vấn" in wb.sheetnames:
            ws_edu = wb["Học vấn"]
            # Xóa học vấn cũ
            db.query(CandidateEducation).filter(CandidateEducation.candidate_id == profile.id).delete()
            # Thêm học vấn mới
            for r in range(2, ws_edu.max_row + 1):
                row_vals = [ws_edu.cell(row=r, column=col).value for col in range(1, 4)]
                if is_instruction_row(row_vals):
                    continue
                    
                inst = ws_edu.cell(row=r, column=1).value
                major = ws_edu.cell(row=r, column=2).value
                deg = ws_edu.cell(row=r, column=3).value
                
                if inst or major or deg:
                    db_edu = CandidateEducation(
                        candidate_id=profile.id,
                        institution_name=str(inst or "").strip(),
                        major=str(major or "").strip(),
                        degree=str(deg or "").strip()
                    )
                    db.add(db_edu)

        # 4. Đọc sheet "Chứng chỉ"
        if "Chứng chỉ" in wb.sheetnames:
            ws_cert = wb["Chứng chỉ"]
            # Xóa chứng chỉ cũ
            db.query(CandidateCertification).filter(CandidateCertification.candidate_id == profile.id).delete()
            # Thêm chứng chỉ mới
            for r in range(2, ws_cert.max_row + 1):
                row_vals = [ws_cert.cell(row=r, column=col).value for col in range(1, 3)]
                if is_instruction_row(row_vals):
                    continue
                    
                cert_name = ws_cert.cell(row=r, column=1).value
                iss = ws_cert.cell(row=r, column=2).value
                
                if cert_name or iss:
                    db_cert = CandidateCertification(
                        candidate_id=profile.id,
                        name=str(cert_name or "").strip(),
                        issuer=str(iss or "").strip()
                    )
                    db.add(db_cert)
                    
        db.commit()
        db.refresh(profile)
        
        # Trả về dữ liệu profile mới nhất sau khi import
        profile_response = CandidateProfileResponse.model_validate(profile)
        return ResponseSchema(
            success=True,
            data=profile_response,
            error=None,
            meta=None
        )
        
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=400, detail=f"Lỗi khi xử lý file Excel: {str(e)}")

